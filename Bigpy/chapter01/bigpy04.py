import pandas as pd
import openpyxl

wb = openpyxl.Workbook()

sheet = wb.active

sheet.title = '회원정보'

#헤더컬럼
header_titles = ['아이디','전화번호']
for idx, title in enumerate(header_titles):
    sheet.cell(row=1,column=idx+1,value=title)


# 내용 저장
members = [('happy','010-1234-5678'),('smile','010-9876-5432')]

row_num = 2

for r, member in enumerate(members) :
    for c,v in enumerate(member) :
        sheet.cell(row=row_num, column=c+1, value= v)
    row_num +=1

wb.save('member.xlsx')
wb.close()